import pandas as pd
import psycopg2
import requests
import logging
from openpyxl import load_workbook
from datetime import datetime

# === ДАННЫЕ ПОДКЛЮЧЕНИЯ (ЗАМЕНИ НА СВОИ ПРИ ИСПОЛЬЗОВАНИИ) ===
conn_params = {
    "host": "your_host",
    "port": 5432,
    "dbname": "your_database",
    "user": "your_user",
    "password": "your_password",
    "sslmode": "require"
}

BOT_TOKEN = "your_bot_token"
CHAT_ID = "your_chat_id"

# === ИМЕНА ФАЙЛОВ ===
date_suffix = datetime.now().strftime("%Y-%m-%d")
output_file = f"channel_id_report_{date_suffix}.xlsx"
log_file = f"report_log_{date_suffix}.log"

# === ЦВЕТА ДЛЯ ВКЛАДОК ===
green_tab = "92D050"   # зелёный
yellow_tab = "FFD966"  # жёлтый
red_tab = "FF0000"     # красный

# === SQL-ЗАПРОСЫ (пример одного канала) ===
queries = {
    "WB ID": """
        WITH wildberries_with_barcode AS (
            SELECT wb.nm_id, m.original_barcode
            FROM stg_wildberries.stg_wildberries_cards_list_main wb
            LEFT JOIN stg_bpe.ms_1c_odata_variants m
                ON wb.nm_id::text = m.id_wildberries::text
        ),
        joined_data AS (
            SELECT
                wbb.nm_id AS channel_id,
                NULLIF(p.id_wildberries, '-') AS "1c_id",
                wbb.original_barcode AS ch_barcode,
                p.products_barcode AS "1c_barcode",
                p.products_brand_name AS brand,
                p.products_article_number AS article,
                p.products_bpe_sku AS sku,
                p.products_name AS name,
                p.products_size AS size,
                CASE 
                    WHEN wbb.nm_id IS NOT NULL AND NULLIF(p.id_wildberries, '-') IS NOT NULL THEN 
                        CASE 
                            WHEN wbb.nm_id::text = NULLIF(p.id_wildberries, '-')::text THEN 'OK'
                            ELSE 'ID MISMATCH'
                        END
                    WHEN wbb.nm_id IS NOT NULL THEN 'ONLY CH'
                    WHEN NULLIF(p.id_wildberries, '-') IS NOT NULL THEN 'ONLY 1C'
                END AS status
            FROM wildberries_with_barcode wbb
            FULL OUTER JOIN bpd.prod_products.products_barcode p
                ON wbb.original_barcode = p.products_barcode
        )
        SELECT * FROM joined_data
        WHERE status != 'OK' AND status IS NOT NULL AND brand = 'Lyle & Scott';
    """
}

# === НАСТРОЙКА ЛОГИРОВАНИЯ ===
logging.basicConfig(
    filename=log_file,
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s"
)

# === ОСНОВНАЯ ЛОГИКА ===
try:
    with psycopg2.connect(**conn_params) as conn:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            for sheet_name, query in queries.items():
                try:
                    df = pd.read_sql_query(query, conn)
                    if df.empty:
                        df = pd.DataFrame({"info": ["Различий в данных нет"]})
                        logging.info(f"{sheet_name}: Нет различий.")
                    else:
                        logging.info(f"{sheet_name}: Обнаружено {len(df)} различий.")
                    df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
                except Exception as e:
                    logging.error(f"{sheet_name}: Ошибка запроса - {str(e)}")
                    pd.DataFrame({"error": [str(e)]}).to_excel(writer, sheet_name=sheet_name[:31], index=False)

    wb = load_workbook(output_file)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cell = ws["A2"].value
        if cell == "Различий в данных нет":
            ws.sheet_properties.tabColor = green_tab
        elif cell == "error":
            ws.sheet_properties.tabColor = red_tab
        else:
            ws.sheet_properties.tabColor = yellow_tab
    wb.save(output_file)
    logging.info("✅ Отчёт сохранён.")

    with open(output_file, "rb") as f:
        response = requests.post(
            f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument",
            data={"chat_id": CHAT_ID},
            files={"document": (output_file, f)}
        )
    if response.status_code == 200:
        logging.info("📤 Файл отправлен в Telegram.")
    else:
        logging.warning(f"⚠️ Ошибка Telegram: {response.text}")

except Exception as e:
    logging.critical(f"❌ Критическая ошибка: {str(e)}")